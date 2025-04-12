import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

# === File Paths ===
grades_file_path = r"i:\Ahmed\websites\python with excel\student.xlsx"
addresses_file_path = r"i:\Ahmed\websites\python with excel\add.xlsx"
merged_output_path = r"i:\Ahmed\websites\python with excel\merged_output.xlsx"

# === Sheet Names ===
grades_sheet_name = "Grades"
addresses_sheet_name = "Sheet1"
output_sheet_name = "Sheet1"

# === Get current year ===
base_year = datetime.now().year

# === Read Excel Sheets ===
grades_df = pd.read_excel(grades_file_path, sheet_name=grades_sheet_name)
addresses_df = pd.read_excel(addresses_file_path, sheet_name=addresses_sheet_name)

# === Merge Data on StudentName ===
merged_df = pd.merge(grades_df, addresses_df, on='StudentName')

# === Generate line codes ===
def generate_line_codes(df, existing=None):
    all_data = df.copy()
    if existing is not None:
        all_data = pd.concat([existing, df], ignore_index=True)

    line_codes = []
    name_counts = {}

    for _, row in df.iterrows():
        name = row['StudentName']
        if name not in name_counts:
            name_counts[name] = existing[existing['StudentName'] == name].shape[0] if existing is not None else 0

        name_counts[name] += 1
        person_id = str(list(name_counts.keys()).index(name) + 1).zfill(3)
        year = base_year + name_counts[name] - 1
        line_code = f"{person_id}-{year}"
        line_codes.append(line_code)

    df['line code'] = line_codes
    return df

# === Process Output File ===
if os.path.exists(merged_output_path):
    try:
        book = load_workbook(merged_output_path)
        if output_sheet_name in book.sheetnames:
            existing_df = pd.read_excel(merged_output_path, sheet_name=output_sheet_name)

            # Generate line codes
            merged_df = generate_line_codes(merged_df, existing=existing_df)

            # Reorder columns to match existing sheet
            ordered_columns = [col for col in existing_df.columns if col in merged_df.columns]
            if 'line code' not in ordered_columns:
                ordered_columns.append('line code')
            merged_df = merged_df[ordered_columns]

            # Write to Excel from second line
            with pd.ExcelWriter(merged_output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                start_row = book[output_sheet_name].max_row
                merged_df.to_excel(writer, sheet_name=output_sheet_name, startrow=start_row, header=False, index=False)

            print("✅ Data appended successfully with line code and dynamic year.")
        else:
            merged_df = generate_line_codes(merged_df)
            merged_df.to_excel(merged_output_path, sheet_name=output_sheet_name, index=False)
            print("✅ Sheet1 created with new data and line codes.")
    except Exception as e:
        print("❌ Error:", e)
else:
    merged_df = generate_line_codes(merged_df)
    merged_df.to_excel(merged_output_path, sheet_name=output_sheet_name, index=False)
    print("✅ New file created and data written with line codes.")